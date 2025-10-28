from ultralytics import YOLO
from collections import deque
from datetime import timedelta
from utils import *
from openpyxl import Workbook, load_workbook
import time
from keras_facenet import FaceNet

attendance_system = None
is_running = False
current_status = "Chưa khởi động"
attendance_log = []
all_students = []


class AttendanceSystem:
    def __init__(self):
        self.cap = None
        self.detect_model = None
        self.embed_model = None
        self.embed_lists = np.empty((0, 512))
        self.name_lists = np.array([])
        self.face_cache = deque(maxlen=50)
        self.attendance_dict = {}
        self.is_running = False
        self.current_frame = None
        self.fps = 0
        self.detected_count = 0
        self.all_students = []

    def initialize_models(self):
        global all_students

        try:
            self.cap = init_camera()
            if not self.cap or not self.cap.isOpened():
                raise Exception("Không thể kết nối camera")

            self.detect_model = YOLO('../outputs/save_model/640_real_fake_2_best.pt')
            self.embed_model = FaceNet()

            if os.path.exists(LIST_EMBEDDING) and os.path.exists(LIST_NAME):
                self.embed_lists = np.load(LIST_EMBEDDING)
                self.name_lists = np.load(LIST_NAME)
                print(f"Đã load {len(self.name_lists)} khuôn mặt từ database")

                self.all_students = list(np.unique(self.name_lists))
                all_students = self.all_students.copy()
                print(f"Danh sách sinh viên: {self.all_students}")

            if self.embed_lists.size > 0:
                self.embed_lists_norm = self.embed_lists / np.linalg.norm(self.embed_lists, axis=1, keepdims=True)

            return True
        except Exception as e:
            print(f"Lỗi khởi tạo: {e}")
            return False

    def start_attendance(self, duration_minutes=30, is_start=False, is_end=False):
        global attendance_log

        if not self.initialize_models():
            return False

        self.is_running = True
        self.attendance_dict = {}
        self.face_cache.clear()
        attendance_log.clear()

        start_time = datetime.now()
        end_time = start_time + timedelta(minutes=duration_minutes)

        path_xlsx = 'diemdanh.xlsx'
        if not os.path.exists(path_xlsx):
            wb = Workbook()
            wb.create_sheet('Sheet1', 0)
            wb.create_sheet('Sheet2', 1)
            ws1 = wb['Sheet1']
            ws2 = wb['Sheet2']
            ws1.append(["Tên"])
            ws2.append(["Tên"])
            for n in np.unique(self.name_lists):
                ws1.append([str(n)])
                ws2.append([str(n)])
            wb.save(path_xlsx)

        THRESHOLD = 0.7
        DETECT_EVERY = 2
        IOU_THRESHOLD = 0.5
        ATTENDANCE_INTERVAL = timedelta(minutes=3)

        frame_count = 0
        last_boxes = []
        fps_counter = 0
        last_fps_time = time.time()
        last_attendance_time = {}

        def iou(boxA, boxB):
            xA, yA = max(boxA[0], boxB[0]), max(boxA[1], boxB[1])
            xB, yB = min(boxA[2], boxB[2]), min(boxA[3], boxB[3])
            inter = max(0, xB - xA) * max(0, yB - yA)
            areaA = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])
            areaB = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])
            return inter / float(areaA + areaB - inter + 1e-6)

        def get_embedding(frame, box):
            x1, y1, x2, y2 = np.clip(np.array(box, int),
                                     [0, 0, 0, 0],
                                     [frame.shape[1] - 1, frame.shape[0] - 1,
                                      frame.shape[1] - 1, frame.shape[0] - 1])
            face = frame[y1:y2, x1:x2]
            if face.size == 0: return None
            face = cv2.resize(face, (160, 160))
            face = np.expand_dims(face, axis=0)
            return self.embed_model.embeddings(face)[0]

        def recognize(embed):
            if self.embed_lists.size == 0:
                return 'Unknown', 0.0
            embed_norm = embed / np.linalg.norm(embed)
            sims = np.dot(self.embed_lists_norm, embed_norm)
            idx = np.argmax(sims)
            score = sims[idx]
            return (str(self.name_lists[idx]), score) if score > THRESHOLD else ("Unknown", score)

        try:
            while self.is_running and datetime.now() < end_time:
                ret, frame = self.cap.read()
                if not ret:
                    break

                frame = cv2.flip(frame, 1)
                frame_count += 1
                fps_counter += 1

                current_time = time.time()
                if current_time - last_fps_time >= 1.0:
                    self.fps = fps_counter
                    fps_counter = 0
                    last_fps_time = current_time

                if frame_count % DETECT_EVERY == 0:
                    results = self.detect_model(frame, conf=0.7, verbose=False)
                    last_boxes = []
                    for r in results:
                        boxes = r.boxes.xyxy.cpu().numpy()
                        classes = r.boxes.cls.cpu().numpy().astype(int)
                        confs = r.boxes.conf.cpu().numpy()
                        for box, cls, conf in zip(boxes, classes, confs):
                            last_boxes.append((box, cls, conf))

                current_datetime = datetime.now()

                for box, cls, conf in last_boxes:
                    x1, y1, x2, y2 = map(int, box)
                    if cls == 1:
                        cached = False
                        name, score = "Unknown", 0.0

                        for c_box, c_embed, c_name, c_score in self.face_cache:
                            if iou((x1, y1, x2, y2), c_box) > IOU_THRESHOLD:
                                name, score = c_name, c_score
                                cached = True
                                break

                        if not cached:
                            embed = get_embedding(frame, (x1, y1, x2, y2))
                            if embed is None:
                                continue
                            name, score = recognize(embed)
                            self.face_cache.append(((x1, y1, x2, y2), embed, name, score))

                        can_attend = (
                                score > THRESHOLD and
                                name != "Unknown" and
                                current_datetime <= end_time
                        )

                        if can_attend:
                            now_str = current_datetime.strftime("%H:%M:%S")

                            if name not in last_attendance_time:
                                self.attendance_dict[name] = now_str
                                last_attendance_time[name] = current_datetime
                                log_entry = f"{name} điểm danh lúc {now_str}"
                                attendance_log.append(log_entry)
                                print(log_entry)

                            elif current_datetime - last_attendance_time[name] >= ATTENDANCE_INTERVAL:
                                old_time = self.attendance_dict.get(name, "")
                                self.attendance_dict[name] = f"{old_time}, {now_str}" if old_time else now_str
                                last_attendance_time[name] = current_datetime
                                log_entry = f"{name} điểm danh lại lúc {now_str}"
                                attendance_log.append(log_entry)
                                print(log_entry)

                        color = (0, 255, 0) if score > THRESHOLD else (0, 0, 255)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

                        status_text = f"{name} {score:.2f}"
                        if name in self.attendance_dict:
                            status_text = f'{name} - like'

                        cv2.putText(frame, status_text, (x1 + 3, y2 - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                        self.detected_count += 1

                    else:
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                        cv2.putText(frame, "fake", (x1 + 3, y1 - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

                self.current_frame = frame

                remaining_time = end_time - current_datetime
                mins, secs = divmod(int(remaining_time.total_seconds()), 60)

                # cv2.putText(frame, f"FPS: {self.fps}", (10, 30),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                # cv2.putText(frame, f"Time: {current_datetime.strftime('%H:%M:%S')}", (10, 60),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                # cv2.putText(frame, f"End: {end_time.strftime('%H:%M')} ({mins:02d}:{secs:02d})", (10, 90),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                # cv2.putText(frame, f"Attendance: {len(self.attendance_dict)}", (10, 120),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        except Exception as e:
            print(f"Lỗi trong quá trình điểm danh: {e}")

        finally:
            if self.attendance_dict:
                self.save_attendance(is_start, is_end, path_xlsx)

            if self.cap:
                self.cap.release()
            self.is_running = False

    def save_attendance(self, is_start, is_end, path_xlsx):
        try:
            wb = load_workbook(path_xlsx)
            ws2 = wb['Sheet2']
            ws1 = wb['Sheet1']

            day_str = f"{datetime.now().day}/{datetime.now().month}"

            if is_start and is_end:
                header = f"{day_str}_end"
            elif is_start:
                header = f"{day_str}_start"
            elif is_end:
                header = f"{day_str}_end"
            else:
                header = f"{day_str}_start"

            headers = [c.value for c in ws2[1]]
            if header not in headers:
                ws2.cell(row=1, column=len(headers) + 1, value=header)
                headers.append(header)
            col_idx = headers.index(header) + 1

            existing_names = [str(c.value).strip() for c in ws2['A'][1:] if c.value]
            for name in self.attendance_dict:
                if name not in existing_names:
                    ws2.append([name])
                    existing_names.append(name)

            name_to_row = {str(ws2.cell(r, 1).value).strip(): r for r in range(2, ws2.max_row + 1)}
            for name, t in self.attendance_dict.items():
                r = name_to_row.get(name)
                if r:
                    ws2.cell(r, col_idx, value=t)

            sheet1_names = [str(c.value).strip() for c in ws1['A'][1:] if c.value]
            for name in existing_names:
                if name not in sheet1_names:
                    ws1.append([name])
                    sheet1_names.append(name)

            day_headers = [c.value for c in ws1[1]][1:]
            if day_str not in day_headers:
                ws1.cell(row=1, column=len(day_headers) + 2, value=day_str)
                day_headers.append(day_str)
            col_day_idx = day_headers.index(day_str) + 2

            for r in range(2, ws2.max_row + 1):
                name = str(ws2.cell(r, 1).value).strip()
                mark = ""
                start_cols = [i + 1 for i, h in enumerate(headers) if
                              h.startswith(f"{day_str}") and h.endswith("_start")]
                for start_col in start_cols:
                    base = headers[start_col - 1].replace("_start", "")
                    end_col = headers.index(f"{base}_end") + 1 if f"{base}_end" in headers else None
                    if end_col:
                        start_val = ws2.cell(r, start_col).value
                        end_val = ws2.cell(r, end_col).value
                        if start_val and end_val:
                            mark = "x"
                            break

                for rr in range(2, ws1.max_row + 1):
                    if str(ws1.cell(rr, 1).value).strip() == name:
                        ws1.cell(rr, col_day_idx, value=mark)
                        break

            wb.save(path_xlsx)
            print(f"Đã cập nhật file {path_xlsx} - {header}")

        except Exception as e:
            print(f"Lỗi khi lưu điểm danh: {e}")

    def stop(self):
        self.is_running = False
        if self.cap:
            self.cap.release()
