import cv2
import os
import sys
import numpy as np
import mediapipe as mp
import math
import tensorflow as tf
import mediapipe as mp
import time
from PIL import Image
from keras_facenet import FaceNet
from sklearn.metrics.pairwise import cosine_similarity
from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


IMAGE_SIZE = 160
DATA_PATH = '../outputs/crop_face_dataset'
LIST_EMBEDDING = '../outputs/save_model/embeds_list.npy'
LIST_NAME = '../outputs/save_model/names_list.npy'
THRESHOLD=0.7

def init_camera(index=1, width=640, height=640):
    cap = cv2.VideoCapture(index)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)

    if not cap.isOpened():
        raise RuntimeError(f"Cannot open camera with index {index}")

    return cap


def euclidean_dist(a, b):
    x1, y1, x2, y2 = a[0], a[1], b[0], b[1]
    return math.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)

def align_face(frame, face, desired_size=(160,160)):
    height, width, _ = frame.shape

    bbox = face.location_data.relative_bounding_box
    x = int(bbox.xmin * width)
    y = int(bbox.ymin * height)
    w = int(bbox.width * width)
    h = int(bbox.height * height)

    pad = 0
    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(width, x + w + pad)
    y2 = min(height, y + h + pad)

    face_roi = frame[y1:y2, x1:x2].copy()
    roi_h, roi_w, _ = face_roi.shape

    left_eye = face.location_data.relative_keypoints[0]
    right_eye = face.location_data.relative_keypoints[1]
    left_eye_pt = (int((left_eye.x*width)-x1), int((left_eye.y*height)-y1))
    right_eye_pt = (int((right_eye.x*width)-x1), int((right_eye.y*height)-y1))

    dy = right_eye_pt[1] - left_eye_pt[1]
    dx = right_eye_pt[0] - left_eye_pt[0]
    angle = np.degrees(np.arctan2(dy, dx))

    eyes_center = ((left_eye_pt[0]+right_eye_pt[0])//2, (left_eye_pt[1]+right_eye_pt[1])//2)
    M = cv2.getRotationMatrix2D(eyes_center, angle, 1.0)
    rotated_face = cv2.warpAffine(face_roi, M, (roi_w, roi_h), flags=cv2.INTER_CUBIC)

    aligned_face = cv2.resize(rotated_face, desired_size)

    return aligned_face

def diemdanh(day, start, end, now, final, name):
    excel_path = 'diem_danh.xlsx'

    if not os.path.exists(excel_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'sheet1'
        ws2 = wb.create_sheet('sheet2')

        ws1.append(['Name'])
        ws2.append(['Name'])

        folder_path = '../outputs/crop_face_dataset'
        if os.path.exists(folder_path):
            names = os.listdir(folder_path)
        else:
            names = []

        for n in names:
            ws1.append([n])
            ws2.append([n])

        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws1 = wb['sheet1']
    ws2 = wb['sheet2']

    headers1 = [cell.value for cell in ws1[1]]
    headers2 = [cell.value for cell in ws2[1]]

    if day not in headers1:
        col_index1 = len(headers1) + 1
        ws1.cell(row=1, column=col_index1, value=day)
        for row in range(2, ws1.max_row + 1):
            ws1.cell(row=row, column=col_index1, value=0)
    else:
        col_index1 = headers1.index(day) + 1

    day_col1 = f"{day}_1"
    day_col2 = f"{day}_2"

    if day_col1 not in headers2:
        ws2.cell(row=1, column=len(headers2) + 1, value=day_col1)
        headers2.append(day_col1)
        for row in range(2, ws2.max_row + 1):
            ws2.cell(row=row, column=len(headers2), value=0)

    if day_col2 not in headers2:
        ws2.cell(row=1, column=len(headers2) + 1, value=day_col2)
        headers2.append(day_col2)
        for row in range(2, ws2.max_row + 1):
            ws2.cell(row=row, column=len(headers2), value=0)

    headers2 = [cell.value for cell in ws2[1]]
    col_day1 = headers2.index(day_col1) + 1
    col_day2 = headers2.index(day_col2) + 1

    name_row = None
    for row in range(2, ws1.max_row + 1):
        if ws1.cell(row=row, column=1).value == name:
            name_row = row
            break

    if name_row is None:
        print(f"Tên '{name}' không tồn tại trong danh sách.")
        return

    current_time_str = datetime.now().strftime('%H:%M:%S')

    if start <= now <= end:
        old_val = ws1.cell(row=name_row, column=col_index1).value
        if old_val is None or old_val == 'x':
            old_val = 0
        elif isinstance(old_val, str) and not old_val.isdigit():
            old_val = 0
        else:
            old_val = int(old_val)

        if not final:
            new_val = old_val + 1
            ws1.cell(row=name_row, column=col_index1).value = new_val
            ws2.cell(row=name_row, column=col_day1).value = current_time_str

        else:
            new_val = old_val + 1
            if new_val > 2:
                new_val = 'x'
            ws1.cell(row=name_row, column=col_index1).value = new_val
            ws2.cell(row=name_row, column=col_day2).value = current_time_str

    else:
        pass

    wb.save(excel_path)
    print(f"✅ Cập nhật điểm danh cho {name} ({day}) thành công!")