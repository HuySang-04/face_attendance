from ultralytics import YOLO
from collections import deque
from datetime import datetime, timedelta
from utils import *
from openpyxl import Workbook, load_workbook

def run_attendance(start_time, end_time, idx, final, path_xlsx='diemdanh.xlsx'):
    cap = init_camera()
    # URL = 'http://192.168.1.30:4747/mjpegfeed?640x640'
    # cap = cv2.VideoCapture(URL)
    detect_model = YOLO('../outputs/save_model/640_real_fake_2_best.pt')
    embed_model = FaceNet()

    try:
        embed_lists = np.load(LIST_EMBEDDING)
        name_lists = np.load(LIST_NAME)
        print(f"Loaded {len(name_lists)} faces from database")
    except Exception as e:
        print(f"Loading error: {e}")
        embed_lists, name_lists = np.empty((0, 512)), np.array([])

    if embed_lists.size > 0:
        embed_lists_norm = embed_lists / np.linalg.norm(embed_lists, axis=1, keepdims=True)

    if not os.path.exists(path_xlsx):
        wb = Workbook()
        wb.create_sheet('Sheet1', 0)
        wb.create_sheet('Sheet2', 1)
        ws1 = wb['Sheet1']
        ws2 = wb['Sheet2']
        ws1.append(["Tên"])
        ws2.append(["Tên"])
        for n in np.unique(name_lists):
            ws1.append([str(n)])
            ws2.append([str(n)])
        wb.save(path_xlsx)
        print(f"Tạo mới file {path_xlsx}")
    else:
        wb = load_workbook(path_xlsx)
        if 'Sheet1' not in wb.sheetnames:
            wb.create_sheet('Sheet1', 0).append(["Tên"])
        if 'Sheet2' not in wb.sheetnames:
            wb.create_sheet('Sheet2', 1).append(["Tên"])
        wb.save(path_xlsx)

    THRESHOLD = 0.7
    DETECT_EVERY = 2
    IOU_THRESHOLD = 0.5
    CACHE_MAXLEN = 50
    face_cache = deque(maxlen=CACHE_MAXLEN)
    attendance_dict = {}
    frame_count = 0
    last_boxes = []
    fps_display = 0
    prev_frame_time = datetime.now().timestamp()

    def iou(boxA, boxB):
        xA, yA = max(boxA[0], boxB[0]), max(boxA[1], boxB[1])
        xB, yB = min(boxA[2], boxB[2]), min(boxA[3], boxB[3])
        inter = max(0, xB - xA) * max(0, yB - yA)
        areaA = (boxA[2]-boxA[0])*(boxA[3]-boxA[1])
        areaB = (boxB[2]-boxB[0])*(boxB[3]-boxB[1])
        return inter / float(areaA + areaB - inter + 1e-6)

    def get_embedding(frame, box):
        x1, y1, x2, y2 = np.clip(np.array(box, int),
                                 [0, 0, 0, 0],
                                 [frame.shape[1]-1, frame.shape[0]-1,
                                  frame.shape[1]-1, frame.shape[0]-1])
        face = frame[y1:y2, x1:x2]
        if face.size == 0: return None
        face = cv2.resize(face, (160, 160))
        face = np.expand_dims(face, axis=0)
        return embed_model.embeddings(face)[0]

    def recognize(embed):
        if embed_lists.size == 0:
            return 'Unknown', 0.0
        embed_norm = embed / np.linalg.norm(embed)
        sims = np.dot(embed_lists_norm, embed_norm)
        idx = np.argmax(sims)
        score = sims[idx]
        return (str(name_lists[idx]), score) if score > THRESHOLD else ("Unknown", score)

    try:
        while True:
            now_time = datetime.now()
            if now_time > end_time:
                print("Hết thời gian điểm danh.")
                break

            ret, frame = cap.read()
            if not ret:
                break

            frame = cv2.flip(frame, 1)
            frame_count += 1

            if frame_count % DETECT_EVERY == 0:
                results = detect_model(frame, conf=0.7, verbose=False)
                last_boxes = []
                for r in results:
                    boxes = r.boxes.xyxy.cpu().numpy()
                    classes = r.boxes.cls.cpu().numpy().astype(int)
                    confs = r.boxes.conf.cpu().numpy()
                    for box, cls, conf in zip(boxes, classes, confs):
                        last_boxes.append((box, cls, conf))

            for box, cls, conf in last_boxes:
                x1, y1, x2, y2 = map(int, box)
                if cls == 1:
                    cached = False
                    name, score = "Unknown", 0.0

                    for c_box, c_embed, c_name, c_score in face_cache:
                        if iou((x1, y1, x2, y2), c_box) > IOU_THRESHOLD:
                            name, score = c_name, c_score
                            cached = True
                            break

                    if not cached:
                        embed = get_embedding(frame, (x1, y1, x2, y2))
                        if embed is None: continue
                        name, score = recognize(embed)
                        face_cache.append(((x1, y1, x2, y2), embed, name, score))

                    if (score > THRESHOLD and name != "Unknown"
                        and name not in attendance_dict
                        and start_time <= now_time <= end_time):
                        attendance_dict[name] = now_time.strftime("%H:%M:%S")
                        print(f"{name} điểm danh lúc {attendance_dict[name]}")

                    color = (0, 255, 0) if score > THRESHOLD else (0, 0, 255)
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                    cv2.putText(frame, f"{name} {score:.2f}", (x1+3, y1-10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                else:
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    cv2.putText(frame, "fake", (x1 + 3, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

            cv2.putText(frame, f"FPS: {fps_display}", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.imshow('Attendance', frame)
            if cv2.waitKey(1) == 27:
                break

    finally:
        cap.release()
        cv2.destroyAllWindows()

        if not attendance_dict:
            print("Không có bản ghi điểm danh để lưu.")
            return

        wb = load_workbook(path_xlsx)
        ws2 = wb['Sheet2']
        ws1 = wb['Sheet1']

        day_str = f"{datetime.now().day}/{datetime.now().month}"
        suffix = f"_{idx}" if idx > 1 else ""
        col_type = "end" if final else "start"
        header = f"{day_str}{suffix}_{col_type}"

        headers = [c.value for c in ws2[1]]
        if header not in headers:
            ws2.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header)
        col_idx = headers.index(header) + 1

        existing_names = [str(c.value).strip() for c in ws2['A'][1:] if c.value]
        for name in attendance_dict:
            if name not in existing_names:
                ws2.append([name])
                existing_names.append(name)

        name_to_row = {str(ws2.cell(r, 1).value).strip(): r for r in range(2, ws2.max_row + 1)}
        for name, t in attendance_dict.items():
            r = name_to_row.get(name)
            if r:
                ws2.cell(r, col_idx, value=t)

        sheet1_names = [str(c.value).strip() for c in ws1['A'][1:] if c.value]
        for name in existing_names:
            if name not in sheet1_names:
                ws1.append([name])
                sheet1_names.append(name)

        day_headers = [c.value for c in ws1[1]][1:]
        if day_str not in day_headers:
            ws1.cell(row=1, column=len(day_headers) + 2, value=day_str)
            day_headers.append(day_str)
        col_day_idx = day_headers.index(day_str) + 2

        for r in range(2, ws2.max_row + 1):
            name = str(ws2.cell(r, 1).value).strip()
            mark = ""
            start_cols = [i + 1 for i, h in enumerate(headers) if h.startswith(f"{day_str}") and h.endswith("_start")]
            for start_col in start_cols:
                base = headers[start_col - 1].replace("_start", "")
                end_col = headers.index(f"{base}_end") + 1 if f"{base}_end" in headers else None
                if end_col:
                    start_val = ws2.cell(r, start_col).value
                    end_val = ws2.cell(r, end_col).value
                    if start_val and end_val:
                        mark = "x"
                        break

            for rr in range(2, ws1.max_row + 1):
                if str(ws1.cell(rr, 1).value).strip() == name:
                    ws1.cell(rr, col_day_idx, value=mark)
                    break

        wb.save(path_xlsx)
        print(f"Đã cập nhật file {path_xlsx} (ca {idx})")

if __name__ =='__main__':
    start = datetime.now()
    end = start + timedelta(minutes=30)
    idx = 2
    run_attendance(start, end, idx, final=False, path_xlsx='../outputs/diemdanh.xlsx')